import datetime as dt
from functools import lru_cache
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config


# --- Chargement et mise en cache du dictionnaire NAF ---
@st.cache_data  # Utiliser le cache Streamlit pour le chargement initial
def load_naf_dictionary(file_path=config.NAF_FILE_PATH):
    """
    Charge le fichier CSV des codes NAF et retourne un dictionnaire Code -> Libellé.
    Tente de lire le fichier avec plusieurs séparateurs et encodages.
    Gère les erreurs de lecture, les fichiers vides, et les doublons de codes.
    """
    # print(f"{dt.datetime.now()} - DEBUG - Loading NAF dictionary from file_path: {file_path}")
    # Attempt to read the CSV with different common separators and encodings.
    attempts = [
        {"sep": ",", "encoding": "utf-8"},
        {"sep": ";", "encoding": "utf-8"},
        {"sep": ",", "encoding": "latin-1"},
        {"sep": ";", "encoding": "latin-1"},
    ]
    df_naf = None
    for attempt in attempts:
        # print(f"{dt.datetime.now()} - DEBUG - Attempting to read NAF CSV with sep='{attempt['sep']}' and encoding='{attempt['encoding']}'")
        try:
            df_naf_current = pd.read_csv(
                file_path,
                sep=attempt["sep"],
                dtype={"Code": str},
                encoding=attempt["encoding"],
            )
            if "Code" in df_naf_current.columns and "Libellé" in df_naf_current.columns:
                # print(f"{dt.datetime.now()} - DEBUG - NAF CSV read successfully with sep='{attempt['sep']}', encoding='{attempt['encoding']}'.")
                df_naf = df_naf_current
                break
            else:
                # print(f"{dt.datetime.now()} - WARNING - NAF CSV read with sep='{attempt['sep']}', encoding='{attempt['encoding']}', but 'Code' or 'Libellé' column missing.")
                pass # Try next attempt
        except (ValueError, pd.errors.ParserError, UnicodeDecodeError) as e:
            # print(f"{dt.datetime.now()} - WARNING - Failed to read NAF CSV with sep='{attempt['sep']}', encoding='{attempt['encoding']}': {e}")
            pass # Try next attempt
        except FileNotFoundError:
            # print(f"{dt.datetime.now()} - ERROR - NAF file not found at path: {file_path}")
            return None
        except pd.errors.EmptyDataError:
            # print(f"{dt.datetime.now()} - ERROR - NAF file is empty at path: {file_path}")
            return None

    if df_naf is None:
        # print(f"{dt.datetime.now()} - ERROR - Could not find 'Code' and 'Libellé' columns in {file_path} with any attempted separator/encoding.")
        return None

    if df_naf.empty:
        # print(f"{dt.datetime.now()} - WARNING - NAF file '{file_path}' is empty or could not be read correctly.")
        return None

    try:
        # Clean column names and 'Code' values.
        df_naf.columns = df_naf.columns.str.strip()
        df_naf["Code"] = df_naf["Code"].astype(str).str.strip()
        # Handle potential duplicate codes by keeping the last occurrence.
        if df_naf["Code"].duplicated().any():
            df_naf = df_naf.drop_duplicates(subset="Code", keep="last")
        naf_dict = df_naf.set_index("Code")["Libellé"].to_dict()
        # print(f"{dt.datetime.now()} - INFO - NAF dictionary loaded successfully with {len(naf_dict)} codes.")
        return naf_dict
    except Exception as e:
        # print(f"{dt.datetime.now()} - ERROR - Critical error during NAF file processing from '{file_path}': {e}")
        return None


# Global variable to store the loaded NAF dictionary.
# Initialized as None and populated by get_naf_lookup().
naf_detailed_lookup = None

def get_naf_lookup():
    """
    Ensures the NAF dictionary is loaded and cached, then populates 
    the global naf_detailed_lookup.
    This function should be called once from the main app after st.set_page_config().
    """
    global naf_detailed_lookup
    # The load_naf_dictionary function is cached by @st.cache_data.
    # This call will either load fresh data or return cached data from Streamlit's cache.
    # The result is assigned to the global naf_detailed_lookup variable for use by other functions in this module.
    if naf_detailed_lookup is None: # Populate only if not already done
        naf_detailed_lookup = load_naf_dictionary()
    return naf_detailed_lookup


# Use lru_cache for functions that are called frequently with the same arguments
# and whose results depend only on those arguments (and the global naf_detailed_lookup).
@lru_cache(maxsize=None)
def get_section_for_code(code):
    if not code or not isinstance(code, str):
        return None
    code_cleaned = code.strip().replace(".", "")[:2]
    section = config.NAF_SECTION_MAP.get(code_cleaned)
    return section


@lru_cache(maxsize=None)
def get_codes_for_section(section_letter):
    """
    Returns a sorted list of NAF codes belonging to a given section letter.
    Relies on the globally loaded naf_detailed_lookup.
    """
    if naf_detailed_lookup is None:
        # print(f"{dt.datetime.now()} - WARNING - get_codes_for_section: NAF dictionary not initialized or failed to load.")
        return []
    if not section_letter:
        return []
    codes = [
        code
        for code in naf_detailed_lookup
        if get_section_for_code(code) == section_letter
    ]
    return sorted(codes)


def correspondance_NAF(code_naf_input):
    """
    Returns the NAF libellé (description) for a given NAF code.
    Handles cases where the code is invalid or not found in the NAF dictionary.
    """
    # print(f"{dt.datetime.now()} - DEBUG - correspondance_NAF called with code_naf_input: '{code_naf_input}'.")
    if naf_detailed_lookup is None:
        # print(f"{dt.datetime.now()} - WARNING - correspondance_NAF: NAF dictionary not initialized or failed to load. Input: '{code_naf_input}'.")
        return f"{code_naf_input} (Dico NAF non chargé)"
    if not code_naf_input or not isinstance(code_naf_input, str):
        # print(f"{dt.datetime.now()} - WARNING - correspondance_NAF: Invalid code_naf_input '{code_naf_input}'.")
        return "Code NAF invalide"
    code_naf_clean = code_naf_input.strip()
    libelle = naf_detailed_lookup.get(code_naf_clean)
    if libelle is None:
        # print(f"{dt.datetime.now()} - WARNING - correspondance_NAF: Code '{code_naf_clean}' not found in NAF dictionary.")
        return f"{code_naf_clean} (Libellé non trouvé)"
    return libelle


def traitement_reponse_api(entreprises, selected_effectifs_codes):
    """
    Processes the raw API response (list of entreprises) into a structured Pandas DataFrame.
    Filters establishments by administrative status ('A' for active) and selected workforce size codes.
    Extracts and combines company and establishment data, including financial information.
    Generates display-friendly columns like NAF descriptions, map colors, and radii.
    """
    # print(f"{dt.datetime.now()} - DEBUG - traitement_reponse_api called. Number of entreprises in input: {len(entreprises) if entreprises else 0}. Selected effectifs codes: {selected_effectifs_codes}")
    if not entreprises:
        return pd.DataFrame()
    all_etablissements_data = []
    processed_sirens = set() # To avoid redundant processing of company-level data for multiple establishments of the same company.
    num_etablissements_processed = 0
    num_etablissements_matched = 0
    for entreprise in entreprises:
        siren = entreprise.get("siren")
        date_creation = entreprise.get("date_creation")
        nombre_etablissements_ouverts = entreprise.get("nombre_etablissements_ouverts")
        code_naf_entreprise = entreprise.get("activite_principale")
        tranche_effectif_salarie_entreprise = entreprise.get("tranche_effectif_salarie")
        tranche_description_entreprise = config.effectifs_tranches.get(
            tranche_effectif_salarie_entreprise, "N/A"
        )
        latest_year_str, ca_latest, resultat_net_latest = None, None, None
        # Process financial data only once per SIREN.
        if siren and siren not in processed_sirens:
            processed_sirens.add(siren)
            finances = entreprise.get("finances", {})
            if finances and isinstance(finances, dict):
                try:
                    available_years = [
                        year for year in finances.keys() if year.isdigit()
                    ]
                    if available_years:
                        latest_year_str = max(available_years)
                        latest_year_data = finances.get(latest_year_str, {})
                        ca_latest = latest_year_data.get("ca")
                        resultat_net_latest = latest_year_data.get("resultat_net")
                except Exception as e:
                    # print(f"{dt.datetime.now()} - WARNING - Error extracting financial data for SIREN {siren}: {e}")
                    latest_year_str = "Erreur"
        matching_etablissements = entreprise.get("matching_etablissements", [])
        for etab in matching_etablissements:
            num_etablissements_processed += 1
            etat_etab = etab.get("etat_administratif")
            tranche_eff_etab = etab.get("tranche_effectif_salarie")
            selected_effectifs_codes_set = (
                set(selected_effectifs_codes)
                if not isinstance(selected_effectifs_codes, set)
                else selected_effectifs_codes
            ) # Ensure selected_effectifs_codes is a set for efficient lookup.
            # Filter for active establishments ('A') matching selected workforce size codes.
            if etat_etab == "A" and tranche_eff_etab in selected_effectifs_codes_set:
                num_etablissements_matched += 1

                # --- Logic for combining Dénomination (nom_complet) and Enseigne ---
                base_name_for_etab = str(entreprise.get("nom_complet", "")).strip() # Company's full name
                if not base_name_for_etab:  # Fallback si nom_complet est vide
                    base_name_for_etab = str(
                        entreprise.get("nom_raison_sociale", "")
                    ).strip()

                liste_enseignes_etab = etab.get("liste_enseignes", [])
                enseignes_str_etab = ""
                if liste_enseignes_etab:
                    # Filter valid 'enseignes' (trade names) and join them.
                    valid_enseignes = [
                        str(e).strip()
                        for e in liste_enseignes_etab
                        if e and str(e).strip()
                    ]
                    if valid_enseignes:
                        enseignes_str_etab = ", ".join(valid_enseignes)

                processed_name = base_name_for_etab

                if enseignes_str_etab and enseignes_str_etab.upper() != "N/A":
                    processed_name_upper = processed_name.upper()
                    enseignes_str_etab_upper = enseignes_str_etab.upper()

                    if not processed_name:
                        # If base name is empty, the 'enseigne' becomes the name.
                        processed_name = enseignes_str_etab
                    # Add 'enseigne' if it's different AND not already a substring of the base name.
                    elif (
                        enseignes_str_etab_upper != processed_name_upper
                        and enseignes_str_etab_upper not in processed_name_upper
                    ):
                        processed_name = f"{processed_name} - {enseignes_str_etab}"

                if not processed_name: # Ensure there's a fallback value if all names are empty.
                    processed_name = "N/A"
                # --- Fin de la logique de combinaison ---

                all_etablissements_data.append(
                    {
                        "SIRET": etab.get("siret"),
                        "SIREN": siren,
                        "tranche_effectif_salarie_etablissement": tranche_eff_etab,
                        "annee_tranche_effectif_salarie": etab.get(
                            "annee_tranche_effectif_salarie"
                        ),
                        "code_naf_etablissement": etab.get("activite_principale"),
                        "adresse": etab.get("adresse"),
                        "latitude": etab.get("latitude"),
                        "Commune": etab.get("libelle_commune"), # <-- ADDED COMMUNE HERE
                        "longitude": etab.get("longitude"),
                        "est_siege": etab.get("est_siege", False),
                        "nom_complet_entreprise": processed_name,  # Use the processed name
                        "nom_sociale_entreprise": entreprise.get(
                            "nom_raison_sociale"
                        ),  # Keep the raw 'raison sociale'
                        "date_creation_entreprise": date_creation,
                        "nb_etab_ouverts_entreprise": nombre_etablissements_ouverts,
                        "code_naf_entreprise": code_naf_entreprise,
                        "tranche_desc_entreprise": tranche_description_entreprise,
                        "annee_finances": latest_year_str,
                        "ca_entreprise": ca_latest,
                        "resultat_net_entreprise": resultat_net_latest,
                    }
                )
    print(
        f"{dt.datetime.now()} - DEBUG - traitement_reponse_api: Processed SIRENs: {len(processed_sirens)}. Total establishments processed: {num_etablissements_processed}. Matched establishments: {num_etablissements_matched}."
    )
    if not all_etablissements_data:
        # print(f"{dt.datetime.now()} - DEBUG - traitement_reponse_api: No establishments matched criteria. Returning empty DataFrame.")
        return pd.DataFrame()
    df_filtered = pd.DataFrame(all_etablissements_data)
    # Add descriptive columns based on codes.
    df_filtered["Activité NAF/APE Entreprise"] = df_filtered[
        "code_naf_entreprise"
    ].apply(lambda x: correspondance_NAF(x) if pd.notna(x) and x != "nan" else "N/A")
    df_filtered["Activité NAF/APE Etablissement"] = df_filtered[
        "code_naf_etablissement"
    ].apply(lambda x: correspondance_NAF(x) if pd.notna(x) and x != "nan" else "N/A")
    df_filtered["Nb salariés établissement"] = df_filtered["tranche_effectif_salarie_etablissement"].map(config.effectifs_tranches).fillna("N/A")
    # Add columns for map visualization.
    df_filtered["Section NAF"] = (
        df_filtered["code_naf_etablissement"].apply(get_section_for_code).fillna("N/A")
    )
    df_filtered["Color"] = df_filtered["Section NAF"].apply(
        lambda section: config.naf_color_mapping.get(
            section, config.naf_color_mapping["N/A"]
        )
    )
    df_filtered["Radius"] = (
        df_filtered["tranche_effectif_salarie_etablissement"]
        .map(config.size_mapping)
        .fillna(config.size_mapping.get("N/A", 10))
    )
    # Convert columns to numeric types, coercing errors.
    df_filtered["Latitude"] = pd.to_numeric(df_filtered["latitude"], errors="coerce")
    df_filtered["Longitude"] = pd.to_numeric(df_filtered["longitude"], errors="coerce")
    df_filtered["Chiffre d'Affaires Entreprise"] = pd.to_numeric(
        df_filtered["ca_entreprise"], errors="coerce"
    )
    df_filtered["Résultat Net Entreprise"] = pd.to_numeric(
        df_filtered["resultat_net_entreprise"], errors="coerce"
    )
    # Rename columns for final output, matching expected export/display names.
    final_df = df_filtered.rename(
        columns={
            "nom_complet_entreprise": "Dénomination - Enseigne",
            "est_siege": "Est siège social",
            "adresse": "Adresse établissement",
            "tranche_effectif_salarie_etablissement": "Code effectif établissement",
            "annee_tranche_effectif_salarie": "Année nb salariés établissement",
            "nom_sociale_entreprise": "Raison sociale",
            "date_creation_entreprise": "Date de création Entreprise",
            "nb_etab_ouverts_entreprise": "Nb total établissements ouverts",
            "tranche_desc_entreprise": "Nb salariés entreprise",
            "annee_finances": "Année Finances Entreprise",
        }
    )
    # Ensure final DataFrame has columns in the order defined in config.COLS_EXPORT_ORDER.
    cols_existantes = [
        col for col in config.COLS_EXPORT_ORDER if col in final_df.columns
    ]
    final_df_result = final_df[cols_existantes]
    # print(f"{dt.datetime.now()} - DEBUG - traitement_reponse_api: Final DataFrame has {len(final_df_result)} rows.")
    return final_df_result

def sanitize_column_name_for_my_maps(name: str) -> str:
    """
    Sanitizes a column name to be compatible with Google My Maps import requirements.
    - Replaces common problematic characters/patterns with underscores.
    - Removes explicitly forbidden characters by Google My Maps.
    - Truncates to 64 characters.
    """
    name = str(name)

    # Replace common problematic patterns first
    name = name.replace(" - ", "_")  # e.g., "Dénomination - Enseigne"
    name = name.replace("/", "_")    # e.g., "NAF/APE"
    
    # Replace remaining spaces with underscores
    name = name.replace(" ", "_")

    # Remove explicitly forbidden characters by Google My Maps: " < > { } |
    forbidden_chars = '"<>{}\\|' # Pipe needs to be escaped for regex, but simple replace is fine.
    for char_to_remove in forbidden_chars:
        name = name.replace(char_to_remove, "")

    # Remove any resulting multiple underscores
    while "__" in name:
        name = name.replace("__", "_")
    
    # Strip leading/trailing underscores that might result
    name = name.strip("_")

    # Truncate to 64 characters
    return name[:64]


def generate_erm_excel(df_entreprises_input: pd.DataFrame):
    """
    Génère un fichier Excel (.xlsx) à partir de zéro avec plusieurs feuilles (DATA_IMPORT, ENTREPRISES, etc.),
    des formules, des données statiques, et des règles de validation de données.

    Args:
        df_entreprises_input (pd.DataFrame): Le DataFrame des résultats de recherche.

    Returns:
        bytes: Le contenu binaire du fichier Excel, ou None en cas d'erreur.
    """
    # print(f"{dt.datetime.now()} - DEBUG - generate_erm_excel called. Input df_entreprises_input shape: {df_entreprises_input.shape}")
    output = BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            workbook = writer.book
            # print(f"{dt.datetime.now()} - DEBUG - ExcelWriter created, workbook obtained.")

            # --- Sheet 1: DATA_IMPORT ---
            # This sheet contains the raw data from the search results, serving as a data source for formulas.
            # print(f"{dt.datetime.now()} - DEBUG - Creating DATA_IMPORT sheet.")
            # Define the columns for the DATA_IMPORT sheet.
            data_import_cols = [
                "SIRET",
                "Dénomination - Enseigne",
                "Activité NAF/APE Etablissement",
                "code_naf_etablissement",
                "Activité NAF/APE Entreprise",
                "code_naf_entreprise",
                "Adresse établissement",
                "Commune", # <-- ADDED COMMUNE
                "Nb salariés établissement",
                "Est siège social",
                "Date de création Entreprise",
                "Chiffre d'Affaires Entreprise",
                "Résultat Net Entreprise",
                "Année Finances Entreprise",
                "SIREN",
            ]
            df_data_import = pd.DataFrame()
            for col in data_import_cols:
                if col in df_entreprises_input.columns:
                    df_data_import[col] = df_entreprises_input[col]
                else:
                    df_data_import[col] = np.nan

            df_data_import = df_data_import[data_import_cols]
            
            # Sanitize column names for DATA_IMPORT sheet
            df_data_import_excel = df_data_import.copy()
            df_data_import_excel.columns = [sanitize_column_name_for_my_maps(col) for col in df_data_import_excel.columns]
            df_data_import_excel.to_excel(
                writer, sheet_name="DATA_IMPORT", index=False, freeze_panes=(1, 0)
            )
            num_data_rows = len(df_data_import) # Based on original data count
            # print(f"{dt.datetime.now()} - DEBUG - DATA_IMPORT sheet created with {num_data_rows} rows.")

            # --- Sheet 2: ENTREPRISES ---
            # This sheet displays processed company information, with some columns populated by formulas referencing DATA_IMPORT.
            # print(f"{dt.datetime.now()} - DEBUG - Creating ENTREPRISES sheet with formulas.")
            entreprises_headers = [
                # No hyperlink columns in the initial header definition
                "SIRET",
                "Dénomination - Enseigne",
                "Recherche LinkedIn",
                "Recherche Google Maps",
                "Recherche Emploi", # Changed from Indeed
                "Activité NAF/APE Etablissement",
                "Commune", # <-- ADDED COMMUNE
                "Adresse établissement",
                "Nb salariés établissement",
                "Est siège social",
                "Date de création Entreprise",
                "Chiffre d'Affaires Entreprise",
                # "Résultat Net Entreprise", # Removed for brevity as in app.py
                # "Année Finances Entreprise", # Removed for brevity as in app.py
            ]
            # Sanitize headers for ENTREPRISES sheet
            sanitized_entreprises_headers = [sanitize_column_name_for_my_maps(h) for h in entreprises_headers]
            df_entreprises_sheet_headers_only = pd.DataFrame(
                columns=sanitized_entreprises_headers
            )
            df_entreprises_sheet_headers_only.to_excel( # Writes sanitized headers
                writer, sheet_name="ENTREPRISES", index=False, freeze_panes=(1, 0)
            )

            ws_entreprises = workbook["ENTREPRISES"]

            # Populate ENTREPRISES sheet with formulas referencing DATA_IMPORT.
            for r_idx in range(num_data_rows):
                excel_row = r_idx + 2  # Excel rows are 1-indexed, data starts on row 2
                
                # Col 1: SIRET
                ws_entreprises.cell(row=excel_row, column=1, value=f"=IF(ISBLANK(DATA_IMPORT!A{excel_row}),\"\",DATA_IMPORT!A{excel_row})")
                # Col 2: Dénomination - Enseigne
                ws_entreprises.cell(row=excel_row, column=2, value=f"=IF(ISBLANK(DATA_IMPORT!B{excel_row}),\"\",DATA_IMPORT!B{excel_row})")
                
                # Col 3: Recherche LinkedIn (references Dénomination - Enseigne (B) and Commune (H) in ENTREPRISES sheet)
                ws_entreprises.cell(
                    row=excel_row,
                    column=3,
                    value=f'=IF(ISBLANK(B{excel_row}),\"\",HYPERLINK("https://www.google.com/search?q="&ENCODEURL(B{excel_row})&"+"&ENCODEURL(H{excel_row})&"+linkedin","Recherche LinkedIn"))'
                )
                # Col 4: Recherche Google Maps (references Dénomination and Adresse établissement in ENTREPRISES sheet)
                ws_entreprises.cell(
                    row=excel_row,
                    column=4,
                    value=f'=IF(OR(ISBLANK(B{excel_row}),ISBLANK(G{excel_row})),\"\",HYPERLINK("https://www.google.com/maps/search/?api=1&query="&ENCODEURL(B{excel_row})&","&ENCODEURL(G{excel_row}),"Recherche Google Maps"))'
                )
                # Col 5: Recherche Emploi (references Dénomination - Enseigne (B) and Commune (H) in ENTREPRISES sheet)
                ws_entreprises.cell(
                    row=excel_row,
                    column=5,
                    value=f'=IF(ISBLANK(B{excel_row}),\"\",HYPERLINK("https://www.google.com/search?q="&ENCODEURL(B{excel_row})&"+"&ENCODEURL(H{excel_row})&"+emploi","Recherche Emploi"))'
                )
                # Col 6: Activité NAF/APE Etablissement
                ws_entreprises.cell(row=excel_row, column=6, value=f"=IF(ISBLANK(DATA_IMPORT!C{excel_row}),\"\",DATA_IMPORT!C{excel_row})")
                # Col 7: Adresse établissement (Source: DATA_IMPORT col G) - this becomes G in ENTREPRISES
                ws_entreprises.cell(row=excel_row, column=7, value=f"=IF(ISBLANK(DATA_IMPORT!G{excel_row}),\"\",DATA_IMPORT!G{excel_row})")
                # Col 8: Commune (Source: DATA_IMPORT col H) - this becomes H in ENTREPRISES
                ws_entreprises.cell(row=excel_row, column=8, value=f"=IF(ISBLANK(DATA_IMPORT!H{excel_row}),\"\",DATA_IMPORT!H{excel_row})")
                # Col 9: Nb salariés établissement (Source: DATA_IMPORT col I)
                ws_entreprises.cell(
                    row=excel_row, column=9, value=f"=IF(ISBLANK(DATA_IMPORT!I{excel_row}),\"\",DATA_IMPORT!I{excel_row})"
                )
                # Col 10: Est siège social (Source: DATA_IMPORT col J)
                ws_entreprises.cell(
                    row=excel_row, column=10, value=f"=IF(ISBLANK(DATA_IMPORT!J{excel_row}),\"\",DATA_IMPORT!J{excel_row})"
                )
                # Col 11: Date de création Entreprise (Source: DATA_IMPORT col K)
                ws_entreprises.cell(
                    row=excel_row, column=11, value=f"=IF(ISBLANK(DATA_IMPORT!K{excel_row}),\"\",DATA_IMPORT!K{excel_row})"
                )
                # Col 12: Chiffre d'Affaires Entreprise (Source: DATA_IMPORT col L)
                ws_entreprises.cell(
                    row=excel_row, column=12, value=f"=IF(ISBLANK(DATA_IMPORT!L{excel_row}),\"\",DATA_IMPORT!L{excel_row})"
                )
                # Col 13: Résultat Net Entreprise (Source: DATA_IMPORT col M)
                ws_entreprises.cell(
                    row=excel_row, column=13, value=f"=IF(ISBLANK(DATA_IMPORT!M{excel_row}),\"\",DATA_IMPORT!M{excel_row})"
                )
                # Col 14: Année Finances Entreprise (Source: DATA_IMPORT col N)
                ws_entreprises.cell(
                    row=excel_row, column=14, value=f"=IF(ISBLANK(DATA_IMPORT!N{excel_row}),\"\",DATA_IMPORT!N{excel_row})"
                )
                # SIREN column (formerly col 14) is removed from ENTREPRISES sheet
            
            # --- Sheet 3: VALEURS_LISTE ---
            # This sheet contains static lists used for data validation dropdowns in other sheets.
            vl_headers = [
                "CONTACTS_Direction",
                "ACTIONS_TypeAction",
                "ACTIONS_StatutAction",
                "ACTIONS_StatutOpportunuiteTaf",
            ]

            vl_data_from_config = {
                "CONTACTS_Direction": config.VALEURS_LISTE_CONTACTS_DIRECTION,
                "ACTIONS_TypeAction": config.VALEURS_LISTE_ACTIONS_TYPEACTION,
                "ACTIONS_StatutAction": config.VALEURS_LISTE_ACTIONS_STATUTACTION,
                "ACTIONS_StatutOpportunuiteTaf": config.VALEURS_LISTE_ACTIONS_STATUTOPPORTUNITE,
            }

            # Pad shorter lists with None to ensure equal length for DataFrame creation.
            max_len = max(len(lst) for lst in vl_data_from_config.values())
            padded_vl_data = {
                col: lst + [None] * (max_len - len(lst))
                for col, lst in vl_data_from_config.items()
            }
            df_valeurs_liste = pd.DataFrame(padded_vl_data)
            df_valeurs_liste = df_valeurs_liste[vl_headers]  # Ensure column order based on original headers

            # Sanitize column names for VALEURS_LISTE sheet
            df_valeurs_liste_excel = df_valeurs_liste.copy()
            df_valeurs_liste_excel.columns = [sanitize_column_name_for_my_maps(col) for col in df_valeurs_liste_excel.columns]
            df_valeurs_liste_excel.to_excel(
                writer, sheet_name="VALEURS_LISTE", index=False, freeze_panes=(1, 0)
            )

            # --- Sheet 4: CONTACTS ---
            # Empty sheet for users to manually input contact information.
            contacts_headers = [
                "Prénom Nom",
                "Entreprise",
                "Poste",
                "Direction",
                "Email",
                "Téléphone",
                "Profil LinkedIn URL",
                "Notes",
            ]
            # Sanitize headers for CONTACTS sheet
            sanitized_contacts_headers = [sanitize_column_name_for_my_maps(h) for h in contacts_headers]
            df_contacts = pd.DataFrame(columns=sanitized_contacts_headers)
            df_contacts.to_excel( # Writes sanitized headers
                writer, sheet_name="CONTACTS", index=False, freeze_panes=(1, 0)
            )

            # --- Sheet 5: ACTIONS ---
            # Empty sheet for users to manually input actions related to contacts/entreprises.
            actions_headers = [
                "Entreprise",
                "Contact (Prénom Nom)",
                "Type Action",
                "Date Action",
                "Description/Notes",
                "Statut Action",
                "Statut Opportunuité Taf",
            ]
            # Sanitize headers for ACTIONS sheet
            sanitized_actions_headers = [sanitize_column_name_for_my_maps(h) for h in actions_headers]
            df_actions = pd.DataFrame(columns=sanitized_actions_headers)
            df_actions.to_excel( # Writes sanitized headers
                writer, sheet_name="ACTIONS", index=False, freeze_panes=(1, 0)
            )

            # --- Data Validation Setup ---
            # print(f"{dt.datetime.now()} - DEBUG - Setting up data validations.")
            max_row_validation = 5000

            # Validations for CONTACTS Sheet
            ws_contacts = workbook["CONTACTS"]
            dv_contacts_entreprise = DataValidation(
                type="list",
                formula1=f"=ENTREPRISES!$B$2:$B${max_row_validation}",
                allow_blank=True,
            )
            dv_contacts_entreprise.error = "Veuillez choisir une entreprise de la liste (Feuille ENTREPRISES, colonne 'Dénomination - Enseigne')."
            dv_contacts_entreprise.errorTitle = "Entreprise Invalide"
            ws_contacts.add_data_validation(dv_contacts_entreprise)
            dv_contacts_entreprise.add(f"B2:B{max_row_validation}")

            dv_contacts_direction = DataValidation(
                type="list",
                formula1=f"=VALEURS_LISTE!$A$2:$A${max_row_validation}",
                allow_blank=True,
            )
            dv_contacts_direction.error = "Veuillez choisir une direction de la liste (Feuille VALEURS_LISTE, colonne 'CONTACTS_Direction')."
            dv_contacts_direction.errorTitle = "Direction Invalide"
            ws_contacts.add_data_validation(dv_contacts_direction)
            dv_contacts_direction.add(f"D2:D{max_row_validation}")

            # Validations for ACTIONS Sheet
            ws_actions = workbook["ACTIONS"]
            dv_actions_entreprise = DataValidation(
                type="list",
                formula1=f"=ENTREPRISES!$B$2:$B${max_row_validation}",
                allow_blank=True,
            )
            dv_actions_entreprise.error = "Veuillez choisir une entreprise de la liste (Feuille ENTREPRISES, colonne 'Dénomination - Enseigne')."
            dv_actions_entreprise.errorTitle = "Entreprise Invalide"
            ws_actions.add_data_validation(dv_actions_entreprise)
            dv_actions_entreprise.add(f"A2:A{max_row_validation}")

            dv_actions_contact = DataValidation(
                type="list",
                formula1=f"=CONTACTS!$A$2:$A${max_row_validation}",
                allow_blank=True,
            )
            dv_actions_contact.error = "Veuillez choisir un contact de la liste (Feuille CONTACTS, colonne 'Prénom Nom')."
            dv_actions_contact.errorTitle = "Contact Invalide"
            ws_actions.add_data_validation(dv_actions_contact)
            dv_actions_contact.add(f"B2:B{max_row_validation}")

            dv_actions_type = DataValidation(
                type="list",
                formula1=f"=VALEURS_LISTE!$B$2:$B${max_row_validation}",
                allow_blank=True,
            )
            dv_actions_type.error = "Veuillez choisir un type d'action (Feuille VALEURS_LISTE, colonne 'ACTIONS_TypeAction')."
            dv_actions_type.errorTitle = "Type d'Action Invalide"
            ws_actions.add_data_validation(dv_actions_type)
            dv_actions_type.add(f"C2:C{max_row_validation}")

            dv_actions_statut = DataValidation(
                type="list",
                formula1=f"=VALEURS_LISTE!$C$2:$C${max_row_validation}",
                allow_blank=True,
            )
            dv_actions_statut.error = "Veuillez choisir un statut d'action (Feuille VALEURS_LISTE, colonne 'ACTIONS_StatutAction')."
            dv_actions_statut.errorTitle = "Statut d'Action Invalide"
            ws_actions.add_data_validation(dv_actions_statut)
            dv_actions_statut.add(f"F2:F{max_row_validation}")

            dv_actions_opportunite = DataValidation(
                type="list",
                formula1=f"=VALEURS_LISTE!$D$2:$D${max_row_validation}",
                allow_blank=True,
            )
            dv_actions_opportunite.error = "Veuillez choisir un statut d'opportunité (Feuille VALEURS_LISTE, colonne 'ACTIONS_StatutOpportunuiteTaf')."
            dv_actions_opportunite.errorTitle = "Statut Opportunité Invalide"
            ws_actions.add_data_validation(dv_actions_opportunite)
            dv_actions_opportunite.add(f"G2:G{max_row_validation}")
            # print(f"{dt.datetime.now()} - DEBUG - Data validations set up.")

            # --- Formatting: Auto-adjust column widths ---
            # print(f"{dt.datetime.now()} - DEBUG - Adjusting column widths.")
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for col_idx, column in enumerate(sheet.columns):
                    max_length = 0
                    column_letter = get_column_letter(col_idx + 1)

                    header_cell = sheet.cell(row=1, column=col_idx + 1)
                    if header_cell.value:
                        max_length = len(str(header_cell.value))

                    # For data-heavy sheets, iterate through some rows to find max content length.
                    if sheet_name in ["DATA_IMPORT", "VALEURS_LISTE"]:
                        for i, cell in enumerate(column):
                            if i == 0:
                                continue
                            try:
                                if cell.value:
                                    cell_len = len(str(cell.value))
                                    if cell_len > max_length:
                                        max_length = cell_len
                            except:
                                pass
                    # For formula-driven sheets, check a few rows or set fixed widths.
                    elif sheet_name == "ENTREPRISES":
                        if col_idx + 1 == 6 or col_idx + 1 == 7:
                            max_length = max(max_length, 30)
                        else:
                            for i in range(min(5, num_data_rows)):
                                cell_value_formula = sheet.cell(
                                    row=i + 2, column=col_idx + 1
                                ).value
                                if cell_value_formula:
                                    pass

                    adjusted_width = (max_length + 2) * 1.2 # Add padding and scaling factor.
                    adjusted_width = min(adjusted_width, 60) # Cap maximum width.
                    sheet.column_dimensions[column_letter].width = adjusted_width

            # --- Styling: Headers and Hyperlinks ---
            header_font = Font(bold=True, color="FFFFFFFF")
            header_fill = PatternFill(
                start_color="FF4472C4", end_color="FF4472C4", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")

            sheets_to_style_headers = [
                "ENTREPRISES",
                "CONTACTS",
                "ACTIONS",
                "VALEURS_LISTE",
            ]
            for sheet_name in sheets_to_style_headers:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for cell in sheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
            
            # Apply hyperlink style to specific columns in the ENTREPRISES sheet.
            ws_entreprises_style = workbook["ENTREPRISES"]
            link_font = Font(color="0563C1", underline="single")
            # Indices de colonnes pour openpyxl (1-basés):
            # Recherche LinkedIn = 3 (C)
            # Recherche Google Maps = 4 (D)
            # Recherche Indeed = 5 (E)
            link_columns_indices = [3, 4, 5]
            for col_idx in link_columns_indices:
                for row_idx in range(2, num_data_rows + 2): # Data rows + header
                    cell = ws_entreprises_style.cell(row=row_idx, column=col_idx)
                    if row_idx > 1: # Skip header
                        cell.font = link_font

            # --- Sheet Management: Hide DATA_IMPORT and Set Order ---
            if "DATA_IMPORT" in workbook.sheetnames:
                ws_data_import = workbook["DATA_IMPORT"]
                ws_data_import.sheet_state = "hidden"

            # Set the desired order of sheets in the workbook.
            # Get all sheet titles currently in the workbook
            current_sheet_titles = [sheet.title for sheet in workbook._sheets]

            # Define the desired visible order
            desired_visible_order = [
                "ENTREPRISES",
                "CONTACTS",
                "ACTIONS",
                "VALEURS_LISTE",
            ]

            final_ordered_sheets = []
            # Add sheets in the desired visible order
            for title in desired_visible_order:
                if title in current_sheet_titles:
                    final_ordered_sheets.append(workbook[title])

            # Add any remaining sheets (like the hidden DATA_IMPORT) to the end of the list.
            for title in current_sheet_titles:
                if title not in desired_visible_order:
                    final_ordered_sheets.append(workbook[title])

            workbook._sheets = final_ordered_sheets
            # print(f"{dt.datetime.now()} - DEBUG - Sheet order set.")

        # print(f"{dt.datetime.now()} - DEBUG - Excel file generation complete. Returning output.")
    except Exception as e:
        # print(f"{dt.datetime.now()} - ERROR - Exception during Excel generation: {e}")
        st.error(
            f"Une erreur est survenue lors de la génération du fichier Excel : {e}"
        )
        import traceback

        st.error(traceback.format_exc())
        return None

    output.seek(0)
    return output.getvalue()


def generate_user_erm_excel(
    df_entreprises: pd.DataFrame, df_contacts: pd.DataFrame, df_actions: pd.DataFrame
) -> bytes:
    """
    Génère un fichier Excel (.xlsx) à partir des DataFrames ERM de l'utilisateur (entreprises, contacts, actions).
    Inclut une feuille DATA_IMPORT (cachée), des feuilles pour chaque type de données ERM,
    des listes de valeurs pour la validation, et des formules de liens hypertextes.
    """
    # print(f"{dt.datetime.now()} - DEBUG - generate_user_erm_excel called.")
    # print(f"{dt.datetime.now()} - DEBUG - Input df_entreprises shape: {df_entreprises.shape}, df_contacts shape: {df_contacts.shape}, df_actions shape: {df_actions.shape}")
    output = BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            workbook = writer.book
            # print(f"{dt.datetime.now()} - DEBUG - ExcelWriter created for user ERM export.")

            # --- Sheet 1: DATA_IMPORT (Hidden) ---
            # Contains a subset of data from the user's entreprises DataFrame, potentially for reference or future use.
            # These are typically columns that might come from an initial API search or import,
            # excluding user-added notes or statuses if they are not part of the "raw" data concept.
            # For this version, we'll include most columns from df_entreprises but ensure
            # it matches the structure expected if it were raw data.
            data_import_cols = [
                "SIRET",
                "Dénomination - Enseigne",
                "Activité NAF/APE Etablissement",
                "Adresse établissement",
                "Commune", # <-- ADDED COMMUNE
                "Nb salariés établissement",
                "Est siège social",
                "Date de création Entreprise",
                "Chiffre d'Affaires Entreprise",
                "Résultat Net Entreprise",
                "Année Finances Entreprise",
                "SIREN",
            ]
            df_data_import = pd.DataFrame()
            for col in data_import_cols:
                if col in df_entreprises.columns:
                    df_data_import[col] = df_entreprises[col]
                else:
                    df_data_import[col] = pd.NA

            df_data_import = df_data_import[data_import_cols]

            # Sanitize column names for DATA_IMPORT sheet
            df_data_import_excel = df_data_import.copy()
            df_data_import_excel.columns = [sanitize_column_name_for_my_maps(col) for col in df_data_import_excel.columns]
            df_data_import_excel.to_excel(
                writer, sheet_name="DATA_IMPORT", index=False, freeze_panes=(1, 0)
            )
            num_data_rows_entreprises = len(
                df_entreprises
            )  # Used for formula ranges later
            
            # --- Sheet 2: ENTREPRISES ---
            # Displays the user's entreprises data, with added hyperlink columns.
            df_entreprises_sheet = df_entreprises.copy()
            num_data_rows_entreprises_sheet = len(df_entreprises_sheet)
            # SIREN, Code effectif établissement, Effectif Numérique are excluded here.
            entreprises_export_base_cols = [
                "SIRET",
                "Dénomination - Enseigne",                
                # No hyperlink columns in the base columns
                # "Recherche LinkedIn",
                # "Recherche Google Maps",
                # "Recherche Emploi",
                "Activité NAF/APE Etablissement",
                "Commune", # <-- ADDED COMMUNE
                "Adresse établissement",
                "Nb salariés établissement",
                "Est siège social",
                "Date de création Entreprise",
                "Chiffre d'Affaires Entreprise",
                # "Résultat Net Entreprise", # Removed for brevity
                # "Année Finances Entreprise", # Removed for brevity
            ]
            
            hyperlink_col_names = ["Recherche LinkedIn", "Recherche Google Maps", "Recherche Emploi"]

            # Build the final list of columns for the ENTREPRISES sheet structure, inserting hyperlink columns.
            final_cols_for_sheet_structure = entreprises_export_base_cols[:]
            denom_index = final_cols_for_sheet_structure.index("Dénomination - Enseigne") + 1
            for col_name in reversed(hyperlink_col_names): # Insert in correct order
                final_cols_for_sheet_structure.insert(denom_index, col_name)

            # Add user-specific columns (e.g., Notes Personnelles) if they exist in the input df_entreprises.
            user_specific_cols_to_check = ["Notes Personnelles", "Statut Piste"]
            for user_col in user_specific_cols_to_check:
                if user_col in df_entreprises.columns:
                    if user_col not in final_cols_for_sheet_structure:
                        final_cols_for_sheet_structure.append(user_col)
            
            # Create a DataFrame with only the necessary columns for the Excel sheet structure
            # Columns for hyperlinks will be present, but their data (formulas) will be written by openpyxl.
            df_entreprises_to_export = pd.DataFrame(columns=final_cols_for_sheet_structure)
            for col in final_cols_for_sheet_structure:
                if col in df_entreprises_sheet.columns and col not in hyperlink_col_names:
                    df_entreprises_to_export[col] = df_entreprises_sheet[col]
                elif col not in hyperlink_col_names: # Ensure all structural columns exist even if not in source df.
                    df_entreprises_to_export[col] = pd.NA
            
            # Ensure correct column order and drop any columns not in final_cols_for_sheet_structure
            df_entreprises_to_export = df_entreprises_to_export.reindex(columns=final_cols_for_sheet_structure)
            
            # Sanitize column names for ENTREPRISES sheet before export
            df_entreprises_to_export_excel = df_entreprises_to_export.copy()
            df_entreprises_to_export_excel.columns = [sanitize_column_name_for_my_maps(col) for col in df_entreprises_to_export_excel.columns]
            df_entreprises_to_export_excel.to_excel(
                writer, sheet_name="ENTREPRISES", index=False, freeze_panes=(1, 0)
            ) # This writes the DataFrame with sanitized headers
            ws_entreprises_export = workbook["ENTREPRISES"]
            
            # Determine column letters for formula references.
            # Assumes "Dénomination - Enseigne" is column B. "Adresse établissement" column letter is found dynamically.
            col_letter_denomination = 'B' # Dénomination - Enseigne
            if "Dénomination - Enseigne" in final_cols_for_sheet_structure:
                 col_letter_denomination = get_column_letter(final_cols_for_sheet_structure.index("Dénomination - Enseigne") + 1)

            adresse_col_name = "Adresse établissement"
            col_letter_adresse = 'A' # Default, will be updated
            if adresse_col_name in final_cols_for_sheet_structure:
                col_letter_adresse = get_column_letter(final_cols_for_sheet_structure.index(adresse_col_name) + 1)

            commune_col_name = "Commune"
            col_letter_commune = 'A' # Default, will be updated
            if commune_col_name in final_cols_for_sheet_structure:
                col_letter_commune = get_column_letter(final_cols_for_sheet_structure.index(commune_col_name) + 1)
            
            # Write HYPERLINK formulas row by row
            for r_idx in range(num_data_rows_entreprises_sheet):
                excel_row = r_idx + 2  # Excel rows are 1-indexed

                # Recherche LinkedIn
                if "Recherche LinkedIn" in final_cols_for_sheet_structure:
                    linkedin_col_idx = final_cols_for_sheet_structure.index("Recherche LinkedIn") + 1
                    # Lire l'URL pré-calculée depuis df_entreprises_sheet (qui est une copie de df_entreprises de app.py)
                    linkedin_url = df_entreprises_sheet.iloc[r_idx]["LinkedIn"] # Utiliser .iloc et le nom de colonne correct
                    if pd.notna(linkedin_url) and str(linkedin_url).strip():
                        linkedin_url_excel = str(linkedin_url).replace('"', '""') # Échapper les guillemets pour Excel
                        ws_entreprises_export.cell(
                            row=excel_row, column=linkedin_col_idx,
                            value=f'=IF(ISBLANK({col_letter_denomination}{excel_row}),"",HYPERLINK("{linkedin_url_excel}","Recherche LinkedIn"))'
                        )
                    else: # URL vide ou NA
                        ws_entreprises_export.cell(
                            row=excel_row, column=linkedin_col_idx,
                            value=f'=IF(ISBLANK({col_letter_denomination}{excel_row}),"","")' # Ou simplement ""
                        )

                # Recherche Google Maps
                if "Recherche Google Maps" in final_cols_for_sheet_structure:
                    gmaps_col_idx = final_cols_for_sheet_structure.index("Recherche Google Maps") + 1
                    gmaps_url = df_entreprises_sheet.iloc[r_idx]["Google Maps"] # Utiliser .iloc et le nom de colonne correct
                    if pd.notna(gmaps_url) and str(gmaps_url).strip():
                        gmaps_url_excel = str(gmaps_url).replace('"', '""')
                        ws_entreprises_export.cell(
                            row=excel_row, column=gmaps_col_idx,
                            value=f'=IF(OR(ISBLANK({col_letter_denomination}{excel_row}),ISBLANK({col_letter_adresse}{excel_row})),"",HYPERLINK("{gmaps_url_excel}","Recherche Google Maps"))'
                        )
                    else: # URL vide ou NA
                        ws_entreprises_export.cell(
                            row=excel_row, column=gmaps_col_idx,
                            value=f'=IF(OR(ISBLANK({col_letter_denomination}{excel_row}),ISBLANK({col_letter_adresse}{excel_row})),"","")' # Ou simplement ""
                        )

                # Recherche Emploi
                if "Recherche Emploi" in final_cols_for_sheet_structure:
                    emploi_col_idx = final_cols_for_sheet_structure.index("Recherche Emploi") + 1
                    emploi_url = df_entreprises_sheet.iloc[r_idx]["Emploi"] # Utiliser .iloc et le nom de colonne correct
                    if pd.notna(emploi_url) and str(emploi_url).strip():
                        emploi_url_excel = str(emploi_url).replace('"', '""')
                        ws_entreprises_export.cell(
                            row=excel_row, column=emploi_col_idx,
                            value=f'=IF(ISBLANK({col_letter_denomination}{excel_row}),"",HYPERLINK("{emploi_url_excel}","Recherche Emploi"))'
                        )
                    else: # URL vide ou NA
                        ws_entreprises_export.cell(
                            row=excel_row, column=emploi_col_idx,
                            value=f'=IF(ISBLANK({col_letter_denomination}{excel_row}),"","")' # Ou simplement ""
                        )
            
            # --- Sheet 3: VALEURS_LISTE ---
            # Contains static lists from config.py for data validation dropdowns.
            vl_headers_base = [ # Use a base list then append
                "CONTACTS_Direction",
                "ACTIONS_TypeAction",
                "ACTIONS_StatutAction",
                "ACTIONS_StatutOpportunuiteTaf",
            ]
            vl_data_from_config_base = {
                "CONTACTS_Direction": config.VALEURS_LISTE_CONTACTS_DIRECTION,
                "ACTIONS_TypeAction": config.VALEURS_LISTE_ACTIONS_TYPEACTION,
                "ACTIONS_StatutAction": config.VALEURS_LISTE_ACTIONS_STATUTACTION,
                "ACTIONS_StatutOpportunuiteTaf": config.VALEURS_LISTE_ACTIONS_STATUTOPPORTUNITE,
            }

            final_vl_headers = list(vl_headers_base) # Make a copy
            final_vl_data_from_config = dict(vl_data_from_config_base) # Make a copy

            # Add "Statut Piste" to VALEURS_LISTE sheet
            statut_piste_header_excel = "Statut Piste" # Column header in Excel as per user request
            if hasattr(config, 'VALEURS_LISTE_ENTREPRISE_STATUTPISTE') and \
               isinstance(config.VALEURS_LISTE_ENTREPRISE_STATUTPISTE, list):
                final_vl_headers.append(statut_piste_header_excel)
                final_vl_data_from_config[statut_piste_header_excel] = config.VALEURS_LISTE_ENTREPRISE_STATUTPISTE
            else:
                # Optional: Log a warning if VALEURS_LISTE_ENTREPRISE_STATUTPISTE is not found or not a list in config
                print(f"Warning: config.VALEURS_LISTE_ENTREPRISE_STATUTPISTE not found or not a list. '{statut_piste_header_excel}' column not added to VALEURS_LISTE sheet.")

            # Pad shorter lists with pd.NA to ensure equal length for DataFrame creation.
            max_len_vl = 0
            if final_vl_data_from_config: # Check if the dictionary is not empty
                # Calculate max length only from actual lists
                valid_lists = [lst for lst in final_vl_data_from_config.values() if isinstance(lst, list)]
                if valid_lists:
                    max_len_vl = max(len(lst) for lst in valid_lists)
            
            padded_vl_data = {
                col_header: (final_vl_data_from_config.get(col_header, []) if isinstance(final_vl_data_from_config.get(col_header), list) else []) + 
                            [pd.NA] * (max_len_vl - len(final_vl_data_from_config.get(col_header, []) if isinstance(final_vl_data_from_config.get(col_header), list) else []))
                for col_header in final_vl_headers
            }

            df_valeurs_liste = pd.DataFrame(padded_vl_data)
            # Ensure column order based on final_vl_headers
            df_valeurs_liste = df_valeurs_liste[final_vl_headers]

            # Sanitize column names for VALEURS_LISTE sheet
            df_valeurs_liste_excel = df_valeurs_liste.copy()
            df_valeurs_liste_excel.columns = [sanitize_column_name_for_my_maps(col) for col in df_valeurs_liste_excel.columns]
            df_valeurs_liste_excel.to_excel(
                writer, sheet_name="VALEURS_LISTE", index=False, freeze_panes=(1, 0)
            )
            
            # --- Sheet 4: CONTACTS ---
            # Displays the user's contacts data.
            df_contacts_sheet = df_contacts.copy()
            expected_contact_cols = [
                "Prénom Nom",
                "Entreprise",
                "Poste",
                "Direction",
                "Email",
                "Téléphone",
                "Profil LinkedIn URL",
                "Notes",
            ]
            for col in expected_contact_cols:
                if col not in df_contacts_sheet.columns:
                    df_contacts_sheet[col] = pd.NA
            df_contacts_sheet = df_contacts_sheet[expected_contact_cols]

            # Sanitize column names for CONTACTS sheet
            df_contacts_sheet_excel = df_contacts_sheet.copy()
            df_contacts_sheet_excel.columns = [sanitize_column_name_for_my_maps(col) for col in df_contacts_sheet_excel.columns]
            df_contacts_sheet_excel.to_excel(
                writer, sheet_name="CONTACTS", index=False, freeze_panes=(1, 0)
            )
            
            # --- Sheet 5: ACTIONS ---
            # Displays the user's actions data.
            df_actions_sheet = df_actions.copy()
            # Ensure 'Date Action' is string or Excel might format it poorly if NaT exists
            if "Date Action" in df_actions_sheet.columns:
                df_actions_sheet["Date Action"] = (
                    df_actions_sheet["Date Action"]
                    .astype(object)
                    .where(df_actions_sheet["Date Action"].notnull(), None)
                )

            expected_action_cols = [
                "Entreprise",
                "Contact (Prénom Nom)",
                "Type Action",
                "Date Action",
                "Description/Notes",
                "Statut Action",
                "Statut Opportunuité Taf",
            ]
            for col in expected_action_cols:
                if col not in df_actions_sheet.columns:
                    df_actions_sheet[col] = pd.NA
            df_actions_sheet = df_actions_sheet[expected_action_cols]

            # Sanitize column names for ACTIONS sheet
            df_actions_sheet_excel = df_actions_sheet.copy()
            df_actions_sheet_excel.columns = [sanitize_column_name_for_my_maps(col) for col in df_actions_sheet_excel.columns]
            df_actions_sheet_excel.to_excel(
                writer, sheet_name="ACTIONS", index=False, freeze_panes=(1, 0)
            )
            
            # --- Data Validation Setup ---
            # print(f"{dt.datetime.now()} - DEBUG - Setting up data validations for user ERM export.")
            max_row_validation = 5000  # Consistent with other function
            # Validations for CONTACTS Sheet
            ws_contacts = workbook["CONTACTS"]
            # Entreprise validation (from ENTREPRISES sheet, 'Dénomination - Enseigne' column - B)
            dv_contacts_entreprise = DataValidation(
                type="list",
                formula1=f"=ENTREPRISES!$B$2:$B${max_row_validation}",
                allow_blank=True,
            )
            ws_contacts.add_data_validation(dv_contacts_entreprise)
            dv_contacts_entreprise.add(f"B2:B{max_row_validation}")
            # Direction validation (from VALEURS_LISTE sheet, column A)
            dv_contacts_direction = DataValidation(
                type="list",
                formula1=f"=VALEURS_LISTE!$A$2:$A${max_row_validation}",
                allow_blank=True,
            )
            ws_contacts.add_data_validation(dv_contacts_direction)
            dv_contacts_direction.add(f"D2:D{max_row_validation}")
            # Validations for ACTIONS Sheet
            ws_actions = workbook["ACTIONS"]
            # Entreprise validation
            dv_actions_entreprise = DataValidation(
                type="list",
                formula1=f"=ENTREPRISES!$B$2:$B${max_row_validation}",
                allow_blank=True,
            )
            ws_actions.add_data_validation(dv_actions_entreprise)
            dv_actions_entreprise.add(f"A2:A{max_row_validation}")
            # Contact validation (from CONTACTS sheet, 'Prénom Nom' column - A)
            dv_actions_contact = DataValidation(
                type="list",
                formula1=f"=CONTACTS!$A$2:$A${max_row_validation}",
                allow_blank=True,
            )
            ws_actions.add_data_validation(dv_actions_contact)
            dv_actions_contact.add(f"B2:B{max_row_validation}")
            # Type Action (from VALEURS_LISTE sheet, column B)
            dv_actions_type = DataValidation(
                type="list",
                formula1=f"=VALEURS_LISTE!$B$2:$B${max_row_validation}",
                allow_blank=True,
            )
            ws_actions.add_data_validation(dv_actions_type)
            dv_actions_type.add(f"C2:C{max_row_validation}")
            # Statut Action (from VALEURS_LISTE sheet, column C)
            dv_actions_statut = DataValidation(
                type="list",
                formula1=f"=VALEURS_LISTE!$C$2:$C${max_row_validation}",
                allow_blank=True,
            )
            ws_actions.add_data_validation(dv_actions_statut)
            dv_actions_statut.add(f"F2:F{max_row_validation}")
            # Statut Opportunuité Taf (from VALEURS_LISTE sheet, column D)
            dv_actions_opportunite = DataValidation(
                type="list",
                formula1=f"=VALEURS_LISTE!$D$2:$D${max_row_validation}",
                allow_blank=True,
            )
            ws_actions.add_data_validation(dv_actions_opportunite)
            dv_actions_opportunite.add(f"G2:G{max_row_validation}")
            # print(f"{dt.datetime.now()} - DEBUG - Data validations set up for user ERM export.")

            # --- Formatting: Auto-adjust column widths ---
            # print(f"{dt.datetime.now()} - DEBUG - Adjusting column widths for user ERM export.")
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for col_idx, column_cells in enumerate(sheet.columns):
                    max_length = 0
                    column_letter = get_column_letter(col_idx + 1)

                    # Header length
                    if sheet.cell(row=1, column=col_idx + 1).value:
                        max_length = len(
                            str(sheet.cell(row=1, column=col_idx + 1).value)
                        )

                    # Cell content length (check first N rows for performance)
                    for i, cell in enumerate(column_cells):
                        if i > 100:
                            break
                        try:
                            if cell.value:
                                cell_len = len(str(cell.value))
                                if cell_len > max_length:
                                    max_length = cell_len
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    adjusted_width = min(adjusted_width, 60)  # Cap max width
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # --- Styling: Headers and Hyperlinks ---
            header_font = Font(bold=True, color="FFFFFFFF")
            header_fill = PatternFill(
                start_color="FF4472C4", end_color="FF4472C4", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")
            sheets_to_style_headers = [
                "ENTREPRISES",
                "CONTACTS",
                "ACTIONS",
                "VALEURS_LISTE",
                "DATA_IMPORT",
            ]
            for sheet_name in sheets_to_style_headers:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for cell in sheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
            
            # Apply hyperlink style to specific columns in the ENTREPRISES sheet.
            ws_entreprises_user_style = workbook["ENTREPRISES"]
            link_font_user = Font(color="0563C1", underline="single")
            # Apply link style to the hyperlink formula columns
            for col_name_link in hyperlink_col_names:
                if col_name_link in final_cols_for_sheet_structure:
                    col_idx_excel = final_cols_for_sheet_structure.index(col_name_link) + 1
                    for row_idx_user in range(2, num_data_rows_entreprises_sheet + 2): # Data rows + header
                        cell_user = ws_entreprises_user_style.cell(row=row_idx_user, column=col_idx_excel)
                        if row_idx_user > 1: # Skip header
                            cell_user.font = link_font_user
            
            # --- Sheet Management: Hide DATA_IMPORT and Set Order ---
            if "DATA_IMPORT" in workbook.sheetnames:
                workbook["DATA_IMPORT"].sheet_state = "hidden"

            desired_visible_order = [
                "ENTREPRISES",
                "CONTACTS",
                "ACTIONS",
                "VALEURS_LISTE",
            ]
            current_sheet_titles = [sheet.title for sheet in workbook._sheets]
            final_ordered_sheets = []
            for title in desired_visible_order:
                if title in current_sheet_titles:
                    final_ordered_sheets.append(workbook[title])
            for title in current_sheet_titles:
                if title not in desired_visible_order:
                    final_ordered_sheets.append(workbook[title])
            workbook._sheets = final_ordered_sheets
            # print(f"{dt.datetime.now()} - DEBUG - Sheet order set for user ERM export.")

    except Exception as e:
        # print(f"{dt.datetime.now()} - ERROR - Exception during user ERM Excel generation: {e}")
        import traceback
        print(traceback.format_exc())
        return None

    output.seek(0)
    # print(f"{dt.datetime.now()} - DEBUG - User ERM Excel file generation complete. Returning output.")
    return output.getvalue()

# The functions add_entreprise_records, ensure_df_schema, and get_erm_data_for_saving
# were part of a previous iteration involving user authentication and are currently not used
# in the global ERM data flow managed directly in app.py.
# They are kept here for potential future re-integration or reference.
def add_entreprise_records(
    current_df_entreprises: pd.DataFrame,
    new_records_df: pd.DataFrame,
    expected_cols: list,
) -> pd.DataFrame:
    """
    Adds new entreprise records to the current DataFrame of entreprises,
    handles deduplication, and ensures schema.
    """
    # print(f"{dt.datetime.now()} - DEBUG - add_entreprise_records called. Current shape: {current_df_entreprises.shape}, New shape: {new_records_df.shape}")


    # Ensure current_df_entreprises has the expected schema
    current_df_processed = current_df_entreprises.copy()
    for col in expected_cols:
        if col not in current_df_processed.columns:
            current_df_processed[col] = pd.NA
    current_df_processed = current_df_processed.reindex(columns=expected_cols)

    # Ensure new_records_df has the expected schema
    new_records_processed = new_records_df.copy()
    for col in expected_cols:
        if col not in new_records_processed.columns:
            new_records_processed[col] = pd.NA
    new_records_processed = new_records_processed.reindex(columns=expected_cols)

    if new_records_processed.empty:
        # print(f"{dt.datetime.now()} - DEBUG - new_records_df is empty. Returning processed current_df_entreprises.")
        return current_df_processed

    # Concatenate
    combined_df = pd.concat(
        [current_df_processed, new_records_processed], ignore_index=True
    )
    # print(f"{dt.datetime.now()} - DEBUG - Shape after concatenation: {combined_df.shape}")
    
    # Drop duplicates, prioritizing the newly added records ('last')
    # Only attempt drop_duplicates if 'SIRET' is present and DataFrame is not empty
    if "SIRET" in combined_df.columns and not combined_df.empty:
        # Fill NA in SIRET column before dropping duplicates to avoid issues with pd.NA comparison if any
        # However, SIRET is expected to be non-null for valid records.
        # If SIRET can be legitimately NA and these rows should be kept, this needs adjustment.
        # For now, assuming SIRET is a key that should exist.
        combined_df.drop_duplicates(subset=["SIRET"], keep="last", inplace=True)
        print(
            f"{dt.datetime.now()} - DEBUG - Shape after dropping duplicates on SIRET: {combined_df.shape}"
        )

    combined_df = combined_df.reindex(columns=expected_cols)
    # print(f"{dt.datetime.now()} - DEBUG - add_entreprise_records finished. Final shape: {combined_df.shape}")
    return combined_df


def ensure_df_schema(df: pd.DataFrame, expected_cols: list) -> pd.DataFrame:
    """
    Ensures the DataFrame has all expected columns with pd.NA for missing ones,
    and reorders columns to match expected_cols.
    """
    # print(f"{dt.datetime.now()} - DEBUG - ensure_df_schema called for DataFrame with shape {df.shape}.")
    df_processed = df.copy()
    added_cols = []
    for col in expected_cols:
        if col not in df_processed.columns:
            df_processed[col] = pd.NA
            added_cols.append(col)
    # if added_cols:
        # print(f"{dt.datetime.now()} - DEBUG - ensure_df_schema: Added missing columns: {added_cols}")
    # else:
        # print(f"{dt.datetime.now()} - DEBUG - ensure_df_schema: No columns were added, all expected columns present.")

    # Ensure correct column order and drop any columns not in expected_cols
    final_df = df_processed.reindex(columns=expected_cols)
    # print(f"{dt.datetime.now()} - DEBUG - ensure_df_schema finished. Output DataFrame shape: {final_df.shape}")
    return final_df
